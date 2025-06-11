from pathlib import Path
import logging
import pandas as pd
from bridgewell_gpt.paths import local_data_path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class BenefitComparisonTemplate:
    """Class for handling benefit comparison Excel template."""
    
    # Define cell mappings for each section
    CELL_MAPPINGS = {
        'life_insurance_ad_d': {
            'schedule': 'B13',
            'reduction': 'B14',
            'non_evidence_maximum': 'B15',
            'termination_age': 'B17'
        },
        'dependent_life': {
            'schedule': 'B21',
            'termination_age': 'B22'
        },
        'critical_illness': {
            'schedule': 'B24',
            'impairments': 'B25',
            'termination_age': 'B35'
        },
        'long_term_disability': {
            'schedule': 'B45',  
            'maximum_monthly_benefit': 'B46',
            'tax_status': 'B47',
            'elimination_period': 'B48',
            'benefit_period': 'B49',
            'definition': 'B50',
            'offsets': 'B51',
            'cost_of_living_adjustment': 'B52',
            'pre_existing': 'B53',
            'survivor_benefit': 'B54',
            'non_evidence_maximum': 'B55',
            'termination_age': 'B56'
        },
        'short_term_disability': {
            'schedule': 'B58',
            'elimination_period': 'B59',
            'maximum_benefit_period': 'B60',
            'maximum_monthly_benefit': 'B61',
            'termination_age': 'B62',
            'non_evidence_maximum': 'B63',
            'termination_age': 'B64'
        },
        'health_care': {
            'prescription_drugs': 'B66',
            'pay_direct_drug_card': 'B69',
            'maximum': 'B70',
            'fertility_drugs': 'B71',
            'smoking_cessations': 'B72',
            'vaccines': 'B73',
            'major_medical': 'B75',
            'annual_deductible': 'B76',
            'hospitalization': 'B77',
            'orthotic_shoes': 'B78',
            'orthotic_inserts': 'B79',
            'hearing_aids': 'B80',
            'vision_care': 'B81',
            'eye_exams': 'B82',
            'paramedical_practitioners': 'B93',
            'included_specialists': 'B95',
            'out_of_country': 'B115',
            'maximum_duration': 'B116',
            'trip_cancellation': 'B121',
            'private_duty_nursing': 'B123',
            'survivor_benefit': 'B125',
            'termination_age': 'B127'
        },
        'dental_care': {
            'annual_deductible': 'B137',
            'basic_preventative': 'B138',
            'periodontic_endodontic': 'B139',
            'annual_maximum': 'B140',
            'major_restorative_services': 'B142',
            'annual_maximum': 'B143',
            'orthodontic_services': 'B146',
            'lifetime_maximum': 'B147',
            'recall_frequency': 'B148',
            'scaling_rooting_units': 'B149',
            'white_filings': 'B151',
            'fee_guide': 'B152',
            'survivor_benefit': 'B153',
            'termination_age': 'B154'
        },
        'notes_and_definitions': {
            'notes': 'B161',
            'definitions': 'B162'
        }
    }
    
    def __init__(self):
        """Initialize the template service with paths."""
        # Set up template and output paths
        self.template_path = local_data_path / "templates" / "benefit_comparison_template.xlsx"
        self.output_dir = local_data_path / "comparisons"
        
        # Create output directory if it doesn't exist
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Create templates directory if it doesn't exist
        self.template_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Check if template exists
        if not self.template_path.exists():
            raise FileNotFoundError(
                f"Template file not found at {self.template_path}. "
                f"Please place the benefit comparison template Excel file at {self.template_path}"
            )
    
    def fill(self, data: dict) -> Path:
        """Fill the template with extracted data.
        
        Args:
            data: Dictionary containing extracted benefit data
            
        Returns:
            Path to the generated Excel file
        """
        output_path = None
        wb = None
        try:
            # Generate output file name
            output_path = self.output_dir / f"benefit_comparison_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Copy template to output location to preserve formatting
            import shutil
            shutil.copy2(self.template_path, output_path)
            
            # Load the workbook with all safety options
            wb = load_workbook(
                filename=str(output_path),
                read_only=False,
                keep_vba=False,
                data_only=True,
                keep_links=False
            )
            ws = wb.active
            
            # Map data to template cells
            for section, fields in self.CELL_MAPPINGS.items():
                if section in data:
                    section_data = data[section]
                    if not isinstance(section_data, dict):
                        logger.warning(f"Skipping section {section} - data is not a dictionary")
                        continue
                        
                    logger.debug(f"Processing section {section} with data: {section_data}")
                    
                    for field, cell in fields.items():
                        if field in section_data:
                            value = section_data[field]
                            # If value is a dict with a 'value' key, use that
                            if isinstance(value, dict) and "value" in value:
                                value = value["value"]
                            value = "" if value is None else str(value).strip()
                            
                            try:
                                # Verify cell exists in worksheet
                                if cell not in ws.merged_cells:  # Skip merged cells
                                    ws[cell] = value
                                    logger.debug(f"Filled cell {cell} with value: {value}")
                            except Exception as cell_error:
                                logger.warning(f"Could not write to cell {cell}: {str(cell_error)}")
                                continue
            
            # Save with proper options
            wb.save(filename=str(output_path))
            wb.close()
            
            # Verify the output file
            try:
                # Try to read with pandas
                pd.read_excel(output_path, engine='openpyxl')
                
                # Try to read with openpyxl again
                verify_wb = load_workbook(filename=str(output_path), read_only=True)
                verify_wb.close()
                
            except Exception as verify_error:
                logger.error(f"Output file verification failed: {str(verify_error)}")
                if output_path.exists():
                    output_path.unlink()
                raise
            
            logger.info(f"Successfully created benefit comparison at {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error filling template: {str(e)}")
            if wb:
                try:
                    wb.close()
                except:
                    pass
            if output_path and output_path.exists():
                output_path.unlink()  # Clean up failed file
            raise
            
        finally:
            # Ensure workbook is closed
            if wb:
                try:
                    wb.close()
                except:
                    pass 